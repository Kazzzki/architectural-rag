import openpyxl
from typing import List, Dict, Any, Tuple

class ExcelAdapter:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=False)
        self.wb_data_only = None # Lazy load

    def _get_data_only_wb(self):
        if not self.wb_data_only:
            self.wb_data_only = openpyxl.load_workbook(self.file_path, data_only=True)
        return self.wb_data_only

    def get_sheets(self) -> List[str]:
        return self.wb.sheetnames

    def get_used_range(self, sheet_name: str) -> str:
        ws = self.wb[sheet_name]
        return ws.dimensions
        
    def get_used_range_indices(self, sheet_name: str) -> Dict[str, List[int]]:
        ws = self.wb[sheet_name]
        return {
            "start": [ws.min_row, ws.min_column],
            "end": [ws.max_row, ws.max_column]
        }
        
    def get_hidden_rows(self, sheet_name: str) -> List[int]:
        ws = self.wb[sheet_name]
        hidden = []
        for row_idx, row_dim in ws.row_dimensions.items():
            if row_dim.hidden:
                hidden.append(row_idx)
        return hidden

    def get_hidden_cols(self, sheet_name: str) -> List[str]:
        ws = self.wb[sheet_name]
        hidden = []
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.hidden:
                hidden.append(col_letter)
        return hidden
        
    def get_formula_regions(self, sheet_name: str) -> List[str]:
        # Highly simplistic heuristic for MVP: search for '=' in the first 1000 rows
        ws = self.wb[sheet_name]
        formula_cells = []
        for row in ws.iter_rows(min_row=1, max_row=min(1000, ws.max_row)):
            for cell in row:
                if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    formula_cells.append(cell.coordinate)
        return formula_cells[:100] # Cap to prevent huge JSON

    def is_sheet_protected(self, sheet_name: str) -> bool:
        ws = self.wb[sheet_name]
        return bool(ws.protection.sheet)

    def get_merged_cells(self, sheet_name: str) -> List[str]:
        ws = self.wb[sheet_name]
        return [str(rng) for rng in ws.merged_cells.ranges]

    def read_range(self, sheet_name: str, min_row: int, max_row: int, min_col: int, max_col: int, data_only: bool = True) -> List[List[Any]]:
        wb = self._get_data_only_wb() if data_only else self.wb
        ws = wb[sheet_name]
        
        result = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            result.append(list(row))
        return result

    def insert_column(self, sheet_name: str, idx: int, header_name: str):
        ws = self.wb[sheet_name]
        ws.insert_cols(idx)
        ws.cell(row=1, column=idx, value=header_name)

    def write_cell(self, sheet_name: str, row: int, col: int, value: Any):
        ws = self.wb[sheet_name]
        ws.cell(row=row, column=col, value=value)

    def write_range(self, sheet_name: str, start_row: int, start_col: int, data: List[List[Any]]):
        ws = self.wb[sheet_name]
        for r_idx, row_data in enumerate(data):
            for c_idx, value in enumerate(row_data):
                ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)

    def set_formula(self, sheet_name: str, row: int, col: int, formula: str):
        ws = self.wb[sheet_name]
        ws.cell(row=row, column=col, value=formula)
        
    def fill_down(self, sheet_name: str, col: int, start_row: int, end_row: int, formula_template: str):
        # simple fill down
        ws = self.wb[sheet_name]
        for r in range(start_row, end_row + 1):
            formula = formula_template.replace("{row}", str(r))
            ws.cell(row=r, column=col, value=formula)

    def set_number_format(self, sheet_name: str, row: int, col: int, number_format: str):
        ws = self.wb[sheet_name]
        ws.cell(row=row, column=col).number_format = number_format

    def add_sheet(self, sheet_name: str):
        if sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(sheet_name)

    def save_as(self, output_path: str):
        self.wb.save(output_path)
