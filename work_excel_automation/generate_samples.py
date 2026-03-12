import openpyxl
import os

os.makedirs("samples", exist_ok=True)
os.makedirs("samples/example_plans", exist_ok=True)

# 1. Sales
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "売上管理"
ws.append(["日付", "案件名", "金額", "担当者", "備考"])
ws.append(["2025-01-01", "システム開発", 1000000, "山田", "新規"])
ws.append(["2025-01-05", "保守運用", 50000, "佐藤", ""])
wb.save("samples/sales_sample.xlsx")

# 2. Project
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Project List"
ws.append(["Project Name", "Assignee", "Amount", "Date", "Status"])
ws.append(["Website Renewal", "Alice", 50000, "2025-02-01", "In Progress"])
ws.append(["App Dev", "Bob", 120000, "2025-03-15", "Planning"])
wb.save("samples/project_sample.xlsx")

# 3. Expense
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "経費精算"
ws.append(["申請日", "摘要", "費用", "申請者", "領収書番号"])
ws.append(["2025-03-01", "交通費", 5000, "田中", "R-001"])
ws.append(["2025-03-02", "書籍代", 3000, "鈴木", "R-002"])
wb.save("samples/expense_sample.xlsx")
